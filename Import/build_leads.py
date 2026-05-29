"""
Подготовка нового варианта импорта клиентов.
Источник: import/Список лидов.xlsx (выгрузка из 1С)
Результат: import/Список лидов — для импорта.xlsx

Правила:
  1) Если ребёнок со статусом "Лид" есть и в других "Состояние лида" — остановка и вывод списка.
  2) Иначе применяется матрица приоритетов:
       а) есть "Черный список" → удалить все остальные записи ребёнка
       б) есть "Архив" (без ЧС)  → удалить все остальные записи ребёнка
       в) одновременно "Потенциал" и "Выбыл" → удалить все "Потенциал"
       г) остальные группы остаются как есть
     После — оставляем одну строку на пару (ФИО, Телефон).
  3) Фамилия родителя приводится к роду по полу родителя (имя из "Контактное лицо"):
     Бочкарёв + Нина → Бочкарёва Нина; Иванова + Сергей → Иванов Сергей.

Итоговые столбцы:
  - Фамилия Имя родителя  (фамилия ребёнка в нужном роде + 1-е слово "Контактное лицо")
  - Номер_телефона        (Телефон)
  - Ребёнок               (ФИО)
  - Соцсети               (Соцсети)
  - Дата_рождения         (Дата рождения)
  - Статус                (Состояние лида)
  - Баланс                (Баланс на сегодня из import/деньги.xlsx, сумма по ребёнку)
  - Проверить             ("да" если автомат не уверен — нужна ручная сверка)
"""

from __future__ import annotations

import sys
import io
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SRC = Path('import/Список лидов.xlsx')
MONEY = Path('import/деньги.xlsx')
DST = Path('import/Список лидов — для импорта.xlsx')

STATE_BLACKLIST = 'Черный список'
STATE_ARCHIVE = 'Архив'
STATE_POTENTIAL = 'Потенциал'
STATE_OUT = 'Выбыл'
STATE_LEAD = 'Лид'


# ---------------------------------------------------------------------------
# Словари имён → пол (для согласования фамилии родителя).
# Имена нормализуются (lower + ё→е) перед сравнением, поэтому в словарь пишем
# одну форму. Покрываем все имена, реально встречающиеся в исходных данных,
# плюс распространённые русские/татарские варианты для устойчивости.
# ---------------------------------------------------------------------------

_F_NAMES = """
александра анастасия анастастия анна анжела анжелика антонина алёна алена алина алия аля
алла алеся альбина альфия альмира альфина анфиса ангелина аделя айгуль айгель айзиря айсель
айсылу айстолу алевтина алиса алися алсу амира амина амелина арзу арина арсения аэлита аида
варвара варя василиса вафа венера вера вероника виктория виолетта вита виля влада валентина
валерия галина галия гелия геллия гелюся гузель гузеля гузял гуля гульназ гульназа гульнара
гульфия гульмира гулина гельназ гельсиня гельфия гильфия глахиаз дана дания дарья дария
джемма джамиля диана дианна дина динара диляра дурсуна ева евгения екатерина елена елизавета
жанна зарима зарина зимфира зинаида зиля зиза зифа зухра зоя зуля ирина илона индира инна
инга ильмира ильвира ильнара ильнура ираида карина катерина катя кадрия камилла клара клавдия
кристина ксения ксюша курбонби лариса лейла лейсан лейсана лейсян ляйсан ляйсян лэйсен лиана
лидия лика лилия лиля лилиля линара леся любовь люба людмила люсьен люсьена люция лерина
мадина маргарита марина мариэта мария марианна марта мелина милена мухалиса наиля настя наталия
наталья натали налья надежда нелли неля нэля нина нонна нозами нурия олеся ольга оксана пикри
полина равия рада раиля раиса рамиля рания рая рейхан регина резеда рина рита роза розалия
роксана румиля румия рузия рузалия рузиля разиля сабина савбагул серафима сирина сириня
сильва снежана софия софья стелла светлана тамара тамила татьяна талина уляна ульзана фания
фая фиюза флюра фруза фатима фидалия фируза хадиджа шахмаза шамиля энже эвелина эльмира
эльвира эльнура эля элеонора эмилия эрна эсмира юлия юля янина яна ягзуль эльза элина залина
"""

_M_NAMES = """
александр алексей анатолий андрей артем артём арсений аждар азат вадим василий виктор виталий
владимир геннадий даниил данис давид денис дмитрий евгений эдуард ильгиз ильдус ильнур ильсур
ильшат игорь леонид линар максим марат михаил николай олег павел петр пётр равиль рамиль
расуль ринат роман руслан рустам рафаэль сергей тимур ярослав юрий
"""

# Уменьшительные и заимствованные с неоднозначным родом — оставляем неуверенными.
_AMBIGUOUS_NAMES = """
саша женя валя миша слава лёня шура карма
"""


def _norm_name_key(s: str) -> str:
    return s.strip().lower().replace('ё', 'е')


FEMALE_NAMES = frozenset(_norm_name_key(n) for n in _F_NAMES.split())
MALE_NAMES = frozenset(_norm_name_key(n) for n in _M_NAMES.split())
AMBIGUOUS_NAMES = frozenset(_norm_name_key(n) for n in _AMBIGUOUS_NAMES.split())


def detect_gender(first_name: str) -> tuple[str | None, bool]:
    """
    Определяет пол по имени.
    Возвращает (пол, уверенно): пол ∈ {'M', 'F', None}, уверенно ∈ {True, False}.
    """
    if not first_name:
        return None, False
    n = _norm_name_key(first_name)
    if not n or n == 'неизвестно':
        return None, False
    # явно не имя — телефон/мусор
    if any(ch.isdigit() for ch in n) or n.startswith('+'):
        return None, False
    if n in FEMALE_NAMES:
        return 'F', True
    if n in MALE_NAMES:
        return 'M', True
    if n in AMBIGUOUS_NAMES:
        return ('F' if n.endswith(('а', 'я')) else 'M'), False
    # неизвестное — по окончанию, неуверенно
    if n.endswith(('а', 'я')):
        return 'F', False
    return 'M', False


# ---------------------------------------------------------------------------
# Согласование рода фамилии.
# ---------------------------------------------------------------------------

_FEMALE_ENDINGS = ('ова', 'ева', 'ёва', 'ина', 'ына', 'ская', 'цкая')
_MALE_ENDINGS = ('ов', 'ев', 'ёв', 'ин', 'ын', 'ский', 'цкий')
_NON_DECLINABLE_ENDINGS = (
    'енко', 'ко', 'ук', 'юк', 'ян', 'швили', 'дзе', 'уа', 'иа',
    'ых', 'их', 'аги', 'оглы',
)


def surname_gender(surname: str) -> str | None:
    """Текущий род фамилии: 'M'/'F' или None если форма нейтральная/несклоняемая."""
    if not surname:
        return None
    s = surname.lower()
    if s.endswith(_NON_DECLINABLE_ENDINGS):
        return None
    if s.endswith(_FEMALE_ENDINGS):
        return 'F'
    if s.endswith(_MALE_ENDINGS):
        return 'M'
    if s.endswith('ская') or s.endswith('цкая'):
        return 'F'
    if s.endswith('ая'):
        return 'F'
    if s.endswith(('ой', 'ый', 'ий')):
        return 'M'
    return None


def feminize(surname: str) -> str | None:
    """Привести мужскую фамилию к женской форме. None — не удалось."""
    s = surname
    sl = s.lower()
    if sl.endswith(_NON_DECLINABLE_ENDINGS):
        return None
    if sl.endswith(_FEMALE_ENDINGS) or sl.endswith('ская') or sl.endswith('цкая') or sl.endswith('ая'):
        return s
    if sl.endswith(('ов', 'ев', 'ёв', 'ин', 'ын')):
        return s + 'а'
    if sl.endswith(('ский', 'цкий', 'ской')):
        return s[:-2] + 'ая'
    if sl.endswith('ой'):
        return s[:-2] + 'ая'
    if sl.endswith(('ый', 'ий')):
        return s[:-2] + 'ая'
    return None


def masculinize(surname: str) -> str | None:
    """Привести женскую фамилию к мужской форме. None — не удалось."""
    s = surname
    sl = s.lower()
    if sl.endswith(_NON_DECLINABLE_ENDINGS):
        return None
    if sl.endswith(_MALE_ENDINGS) or sl.endswith(('ой', 'ый', 'ий')):
        return s
    if sl.endswith(('ова', 'ева', 'ёва', 'ина', 'ына')):
        return s[:-1]
    if sl.endswith('ская'):
        return s[:-2] + 'ий'
    if sl.endswith('цкая'):
        return s[:-2] + 'ий'
    # «-ая» неоднозначно (Толстая→Толстой, Белая→Белый) — пропускаем.
    return None


def align_surname(child_surname: str, parent_gender: str | None) -> tuple[str, bool]:
    """
    Возвращает (фамилия для родителя, ok).
    ok=False — преобразование сорвалось или нужна сверка.
    """
    if not child_surname:
        return child_surname, True
    if parent_gender is None:
        return child_surname, False
    current = surname_gender(child_surname)
    if current is None:
        # форма нейтральная/несклоняемая — оставляем как есть; это не ошибка
        return child_surname, True
    if current == parent_gender:
        return child_surname, True
    transformed = feminize(child_surname) if parent_gender == 'F' else masculinize(child_surname)
    if transformed is None:
        return child_surname, False
    return transformed, True


def norm_name(x) -> str:
    if pd.isna(x):
        return ''
    return ' '.join(str(x).strip().lower().split())


def norm_phone(x):
    if pd.isna(x):
        return None
    s = str(x)
    if s.endswith('.0'):
        s = s[:-2]
    digits = ''.join(c for c in s if c.isdigit())
    return digits or None


def fmt_phone(x):
    if pd.isna(x):
        return ''
    s = str(x)
    if s.endswith('.0'):
        s = s[:-2]
    digits = ''.join(c for c in s if c.isdigit())
    return digits


def fmt_date(x):
    if pd.isna(x) or x == '' or x is None:
        return ''
    if isinstance(x, pd.Timestamp):
        return x.strftime('%d.%m.%Y')
    s = str(x).strip()
    # уже строка в нужном формате — оставляем
    return s


def first_word(s) -> str:
    if pd.isna(s):
        return ''
    parts = str(s).strip().split()
    return parts[0] if parts else ''


def parent_full_name(child_fio, contact_person) -> tuple[str, bool, bool]:
    """
    Возвращает (фамилия+имя родителя, нужна_проверка, фамилия_была_изменена).
    Логика:
      - 1-е слово ФИО ребёнка = база фамилии;
      - 1-е слово «Контактное лицо» = имя родителя;
      - по имени родителя определяем пол;
      - подгоняем род фамилии под родителя.
    """
    base_surname = first_word(child_fio)
    parent_name = first_word(contact_person)
    if not base_surname and not parent_name:
        return '', True, False  # пусто и подозрительно
    gender, gender_confident = detect_gender(parent_name)
    aligned, align_ok = align_surname(base_surname, gender)
    changed = (aligned != base_surname)
    # «Проверить» = да, если:
    #   - пол не определён вовсе, ИЛИ
    #   - пол определён неуверенно и мы поменяли фамилию, ИЛИ
    #   - попытались поменять, но не смогли (несклоняемая в нестандартной форме).
    needs_review = False
    if gender is None:
        needs_review = True
    elif not align_ok:
        needs_review = True
    elif changed and not gender_confident:
        needs_review = True
    full = f'{aligned} {parent_name}'.strip()
    return full, needs_review, changed


def load_balances() -> dict[str, float]:
    """Читает деньги.xlsx и возвращает {нормализованное ФИО → сумма «Баланс на сегодня»}."""
    if not MONEY.exists():
        print(f'⚠ Файл {MONEY} не найден — баланс не загружен.')
        return {}
    m = pd.read_excel(MONEY, header=0)
    m = m[m['Контрагент'].astype(str).str.strip().str.lower() != 'итого']
    m = m.dropna(subset=['Контрагент'])
    m['_key'] = m['Контрагент'].apply(norm_name)
    m['Баланс на сегодня'] = pd.to_numeric(m['Баланс на сегодня'], errors='coerce').fillna(0)
    grouped = m.groupby('_key')['Баланс на сегодня'].sum()
    return grouped.to_dict()


def main() -> int:
    df = pd.read_excel(SRC, header=3)
    df['_row'] = range(len(df))
    df['_name'] = df['ФИО'].apply(norm_name)
    df['_phone'] = df['Телефон'].apply(norm_phone)
    balances = load_balances()
    print(f'✔ Загружено балансов из деньги.xlsx: {len(balances)} уникальных контрагентов.')

    def make_key(row):
        if row['_phone'] is None:
            return f'__solo_{row["_row"]}'
        return f'{row["_name"]}|{row["_phone"]}'

    df['_key'] = df.apply(make_key, axis=1)

    # --- Правило 1: Лид не может соседствовать с другими статусами ---
    #   (а) у одной пары (ФИО, Телефон): Лид + другой статус;
    #   (б) у одного телефона (= родителя): Лид-ребёнок и не-Лид-ребёнок.
    lid_keys = set(df.loc[df['Состояние лида'] == STATE_LEAD, '_key'])
    conflicts = []
    # (а) пара (ФИО+Телефон)
    for key in lid_keys:
        g = df[df['_key'] == key]
        other_states = set(g['Состояние лида'].dropna()) - {STATE_LEAD}
        if other_states:
            row = g.iloc[0]
            conflicts.append({
                'ФИО': row['ФИО'],
                'Телефон': fmt_phone(row['Телефон']),
                'Состояния': ', '.join(sorted(set(g['Состояние лида'].dropna()))),
                'Тип': 'один ребёнок в Лиде и другом статусе',
            })
    # (б) телефон (родитель)
    lid_phones = set(df.loc[(df['Состояние лида'] == STATE_LEAD) & (df['_phone'].notna()), '_phone'])
    for phone in lid_phones:
        g = df[df['_phone'] == phone]
        states = set(g['Состояние лида'].dropna())
        if STATE_LEAD in states and (states - {STATE_LEAD}):
            for _, row in g.iterrows():
                conflicts.append({
                    'ФИО': row['ФИО'],
                    'Телефон': fmt_phone(row['Телефон']),
                    'Состояния': row['Состояние лида'] if pd.notna(row['Состояние лида']) else '',
                    'Тип': 'на одном телефоне дети в Лиде и других статусах',
                })

    if conflicts:
        print('⛔ Остановка по правилу 1. Конфликты "Лид + другой статус":')
        for c in conflicts:
            print(f'  - {c["ФИО"]} | {c["Телефон"]} | {c["Состояния"]} | {c["Тип"]}')
        return 1

    print(f'✔ Правило 1: конфликтов не найдено (Лид-записей: {len(lid_keys)}).')

    # --- Правило 2: матрица приоритетов ---
    def apply_priority(g: pd.DataFrame) -> pd.DataFrame:
        states = set(g['Состояние лида'].dropna())
        if STATE_BLACKLIST in states:
            return g[g['Состояние лида'] == STATE_BLACKLIST]
        if STATE_ARCHIVE in states:
            return g[g['Состояние лида'] == STATE_ARCHIVE]
        if STATE_POTENTIAL in states and STATE_OUT in states:
            return g[g['Состояние лида'] == STATE_OUT]
        return g

    parts, changed = [], 0
    for _, g in df.groupby('_key', sort=False):
        f = apply_priority(g)
        if len(f) != len(g):
            changed += 1
        parts.append(f)
    filtered = pd.concat(parts, ignore_index=True)
    print(f'✔ Правило 2: затронуто групп — {changed}; строк осталось {len(filtered)} из {len(df)}.')

    # --- Дедупликация: одна строка на пару (ФИО, Телефон) ---
    out_rows = []
    surname_changed = 0
    surname_review = 0
    for key, g in filtered.groupby('_key', sort=False):
        g = g.reset_index(drop=True)
        base = g.iloc[0]
        socs = g['Соцсети'].dropna().astype(str).map(str.strip)
        socs = [s for s in socs if s]
        soc_value = '; '.join(dict.fromkeys(socs)) if socs else ''
        contact = next((str(x).strip() for x in g['Контактное лицо'] if not pd.isna(x) and str(x).strip()), '')
        dob = next((x for x in g['Дата рождения'] if not pd.isna(x) and str(x).strip() != ''), None)

        parent, needs_review, changed = parent_full_name(base['ФИО'], contact)
        if changed:
            surname_changed += 1
        if needs_review:
            surname_review += 1

        child_name = str(base['ФИО']).strip() if not pd.isna(base['ФИО']) else ''
        balance = balances.get(norm_name(child_name))

        out_rows.append({
            'Фамилия Имя родителя': parent,
            'Номер_телефона': fmt_phone(base['Телефон']),
            'Ребёнок': child_name,
            'Соцсети': soc_value,
            'Дата_рождения': fmt_date(dob),
            'Статус': str(base['Состояние лида']).strip() if not pd.isna(base['Состояние лида']) else '',
            'Баланс': balance if balance is not None else '',
            'Проверить': 'да' if needs_review else '',
        })

    out_df = pd.DataFrame(out_rows, columns=[
        'Фамилия Имя родителя', 'Номер_телефона', 'Ребёнок',
        'Соцсети', 'Дата_рождения', 'Статус', 'Баланс', 'Проверить',
    ])
    matched_balances = sum(1 for r in out_rows if r['Баланс'] != '')
    print(f'✔ Баланс проставлен у {matched_balances} из {len(out_df)} строк.')
    print(f'✔ Итоговая таблица: {len(out_df)} строк.')
    print(f'✔ Фамилия родителя приведена к роду: изменено {surname_changed} строк.')
    print(f'✔ На ручную проверку («Проверить = да»): {surname_review} строк.')
    print('\nРаспределение по статусам:')
    print(out_df['Статус'].value_counts(dropna=False).to_string())

    # --- Запись в xlsx с оформлением ---
    wb = Workbook()
    ws = wb.active
    ws.title = 'Лиды'

    headers = list(out_df.columns)
    ws.append(headers)
    for _, r in out_df.iterrows():
        ws.append([r[h] for h in headers])

    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
    body_font = Font(name='Arial')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for col_idx, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    phone_col_idx = headers.index('Номер_телефона') + 1
    balance_col_idx = headers.index('Баланс') + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for c in row:
            c.font = body_font
            c.alignment = left
        ph = row[phone_col_idx - 1]
        ph.number_format = '@'
        if ph.value is not None:
            ph.value = str(ph.value)
        bal = row[balance_col_idx - 1]
        bal.number_format = '#,##0;-#,##0;-'
        bal.alignment = Alignment(horizontal='right', vertical='center')

    widths = {
        'Фамилия Имя родителя': 32,
        'Номер_телефона': 18,
        'Ребёнок': 32,
        'Соцсети': 40,
        'Дата_рождения': 16,
        'Статус': 18,
        'Баланс': 14,
        'Проверить': 12,
    }
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths[h]

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{ws.max_row}'
    ws.row_dimensions[1].height = 28

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)
    print(f'\n✔ Сохранено: {DST}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
